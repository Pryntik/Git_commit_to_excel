import re
import openpyxl
import os
import pandas as pd
from io import StringIO
from datetime import datetime
from openpyxl.styles import PatternFill, Alignment

# Headers
H_TITLE = 'Title'
H_DESCRIPTION = 'Description'
H_COMMIT_ID = 'Commit ID'
H_AUTHOR = 'Author'

# Attributes
A_TYPE = 'Type'
A_YEAR = 'Year'
A_MONTH = 'Month'

# Constants
HEADER_KEYS = [H_TITLE, H_DESCRIPTION, H_COMMIT_ID, H_AUTHOR]
ATTRIBUTES_KEYS = [A_TYPE, A_YEAR, A_MONTH]
ATTRIBUTES_COLORED = [A_YEAR, A_MONTH]
BACKGROUND_HEXA_COLORS = ['FFE599', 'FFF2CC']
COLUMNS_EMPTY_FOR_SEPARATOR = [H_DESCRIPTION, H_COMMIT_ID, H_AUTHOR]
ROW_EMPTY_VALUES = {column: '' for column in COLUMNS_EMPTY_FOR_SEPARATOR}

# Dictionaries
MONTH_SHORT_TO_LONG = {
    'Jan': 'January', 'Feb': 'February', 'Mar': 'March', 'Apr': 'April',
    'May': 'May', 'Jun': 'June', 'Jul': 'July', 'Aug': 'August',
    'Sep': 'September', 'Oct': 'October', 'Nov': 'November', 'Dec': 'December'
}
MONTH_SHORT_TO_NUM = {
    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
    'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
}

# Assets
root_dir = os.path.dirname(os.path.abspath(__file__))
asset_dir = os.path.join(root_dir, 'assets')
output_dir = os.path.join(root_dir, 'output')
os.makedirs(asset_dir, exist_ok=True)
os.makedirs(output_dir, exist_ok=True)

# Files
file_name = 'commits'
file_input = f'{file_name}.txt'
file_output = f'{file_name}.xlsx'

# Paths
path_input = os.path.join(asset_dir, file_input)
path_output =  os.path.join(output_dir, file_output)

# Styles
left_alignment = Alignment(horizontal='left', vertical='top')
column_widths = {
    H_TITLE: 100,
    H_DESCRIPTION: 150
    # Not defined = fit-content
}
colors_separator = {
    attribute: PatternFill(start_color=color, end_color=color, fill_type="solid")
    for attribute, color in zip(ATTRIBUTES_COLORED, BACKGROUND_HEXA_COLORS)
}

# Console command to get your git log (put it in the assets/commits.txt file) :
# git log --all --decorate --graph
with open(path_input, encoding='utf-8') as f:
    commit_text = f.read()

# Commits variables
commit_blocks = re.split(r"^\s*[\s|/\\]*\*\s*[\s|/\\]*commit\s+", commit_text, flags=re.MULTILINE)[1:]
commits_datas = {}

# Date variables
year_val = 0
month_val = 0
year_str = ''
month_str = ''

for block in commit_blocks:
    block_lines = block.strip().split('\n')
    lines = []
    for current_line in block_lines:
        # Remove all '|', '\', '/' at start and strip the line
        cleaned_line = re.sub(r"^[|\s/\\]*", "", current_line).strip()
        if cleaned_line:
            lines.append(cleaned_line)

    commit_id = lines[0] if lines else ''

    # Author
    author_line = next((line for line in lines if line.startswith('Author:')), '')
    author_match = re.search(r'Author: (.+?) <(.+?)>', author_line)
    if author_match:
        author_name = author_match.group(1)
        author_email = author_match.group(2)
    else:
        author_name, author_email = '', ''

    # Date
    date_line = next((line for line in lines if line.startswith('Date:')), '')
    date_match = re.search(r'(\w{3}) (\w{3}) (\d{1,2}) (\d{2}:\d{2}:\d{2}) (\d{4})(?: ([\+\-]\d{4}))?', date_line)
    if date_match:
        year_str = date_match.group(5)
        month_short = date_match.group(2)

        year_val = int(year_str)
        month_val = MONTH_SHORT_TO_NUM.get(month_short, 0)
        month_str = MONTH_SHORT_TO_LONG.get(month_short)
    else:
        year_str, month_str = '', ''

    message_lines = [
        line for i, line in enumerate(lines)
        if i != 0 and not (line.startswith('Author:') or line.startswith('Date:')) and line
    ]

    # Title and description
    title = message_lines[0] if message_lines else ''
    description = '\n'.join(message_lines[1:]) if len(message_lines) > 1 else ''

    # Commits associated with a year and a month
    key_tuple = (year_val, month_val)
    commits_datas.setdefault(key_tuple, []).append({
        A_TYPE: 'Commit',
        A_YEAR: year_val,
        A_MONTH: month_str,
        H_TITLE: title,
        H_DESCRIPTION: description,
        H_COMMIT_ID: commit_id,
        H_AUTHOR: author_name,
    })

# Generating rows for the xlsx table
rows = []
previous_year, previous_month = None, None
sorted_commits_keys = sorted(commits_datas.keys(), reverse=True)

for key in sorted_commits_keys:
    # Check if the commit with key is valid
    commit_group = commits_datas[key]
    if not commit_group:
        continue

    current_year = commit_group[0][A_YEAR]
    current_month = commit_group[0][A_MONTH]

    if current_year != previous_year:
        year_row = {A_TYPE: A_YEAR, H_TITLE: current_year}
        year_row.update(ROW_EMPTY_VALUES)
        rows.append(year_row)
        previous_year = current_year
        previous_month = None

    if current_month != previous_month:
        month_row = {A_TYPE: A_MONTH, H_TITLE: current_month}
        month_row.update(ROW_EMPTY_VALUES)
        rows.append(month_row)
        previous_month = current_month

    for commit in commit_group:
        rows.append({
            H_TITLE: commit[H_TITLE],
            H_COMMIT_ID: commit[H_COMMIT_ID],
            H_DESCRIPTION: commit[H_DESCRIPTION],
            H_AUTHOR: commit[H_AUTHOR]
        })

# Backup in an xlsx file
allDataFrame = pd.DataFrame(rows)
dataFrame = allDataFrame.drop(columns=[A_TYPE], errors='ignore')
dataFrame.to_excel(path_output, index=False)

# Adjust the width of each column
workbook = openpyxl.load_workbook(path_output)
ws = workbook.active
for i, column_cells in enumerate(ws.columns):
    column_letter = column_cells[0].column_letter
    header = ws.cell(row=1, column=column_cells[0].column).value

    if header in column_widths:
        # Custom width
        ws.column_dimensions[column_letter].width = column_widths[header]
    else:
        # Fit-content
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_letter].width = length + 2 if length > 0 else length

# Apply background color to the date separator rows
for row_id, row_data in enumerate(allDataFrame.to_dict('records'), start=2):
    is_separator_row = all(not row_data.get(col_header) for col_header in COLUMNS_EMPTY_FOR_SEPARATOR)
    if is_separator_row:
        for col_id in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_id, column=col_id)
            if row_data[A_TYPE] in colors_separator:
                cell.fill = colors_separator[row_data[A_TYPE]]

# Apply alignment to cells
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = left_alignment

# Save the modified file
workbook.save(path_output)
print(f"\033[92mfile xlsx generated successfully\n\033[0mlocation : {path_output}")